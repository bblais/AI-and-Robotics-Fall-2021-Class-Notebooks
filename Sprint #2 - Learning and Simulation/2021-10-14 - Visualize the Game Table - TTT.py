#!/usr/bin/env python
# coding: utf-8

# In[1]:


from Game import *


# In[2]:


def initial_state():
    state=Board(3,3)
    state.pieces=[".","X","O"]
    return state

def show_state(state):
    print(state)
    
def valid_moves(state,player):
    # run through all the spots
    # if it is empty, then append that
    # location to the possible moves
    
    moves=[]
    for location in range(9):
        if state[location]==0:
            moves.append(location)
            
    return moves  

def update_state(state,player,move):
    new_state=state
    
    new_state[move]=player
    return new_state    
    
def win_status(state,player):
    # the state is *after* the move for the player

    #  0  1  2 
    #  3  4  5 
    #  6  7  8   
    
    for start,middle,end in [
        [0,1,2],[3,4,5],[6,7,8],
        [0,3,6],[1,4,7],[2,5,8],
        [0,4,8],[2,4,6],
                ]:
        
        if state[start]==player and state[middle]==player and state[end]==player:
            return "win"
        
    if player==1:
        other_player=2
    else:
        other_player=1
    
    if not valid_moves(state,other_player):
        return "stalemate"
    
    
    
    


# In[63]:


try:
    import openpyxl
except ModuleNotFoundError:
    get_ipython().system('pip install openpyxl')


# In[64]:


from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Color,PatternFill,Border,Side


# In[73]:


def format_cell(cell,color):
    
    cell.alignment=Alignment(horizontal='center',wrapText=True)
    cell.fill=PatternFill(patternType='solid',fgColor=color)  
    
    cell.border=Border(left=Side(border_style='thick',
                           color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),)

def save_table(T,filename):
    wb=Workbook()
    ws=wb.active

    cell=ws.cell(column=1,row=1,value="State")
    cell.font=Font(bold=True)
    cell.alignment=Alignment(horizontal='center')

    cell=ws.cell(column=2,row=1,value="Action")
    cell.font=Font(bold=True)
    cell.alignment=Alignment(horizontal='center')

    cell=ws.cell(column=3,row=1,value="Value")
    cell.font=Font(bold=True)
    cell.alignment=Alignment(horizontal='center')

    # color codes from https://www.w3schools.com/colors/colors_picker.asp
    colors=['#99ccff','#ffff99']  # cycle through these colors per state.  
    color_count=0

    row=2
    for state_as_list in T:
        state=initial_state()
        state.board=state_as_list
        state_str=str(state).strip()

        color=colors[color_count].replace('#','')
        color_count+=1
        if color_count>=len(colors):
            color_count=0

        print(state_str)
        for a,action in enumerate(T[state]):
            value=T[state][action]
            print("\t",action,"\t",value)

            for col,val in zip([1,2,3],[state_str,str(action),value]):
                cell=ws.cell(column=col,row=row,value=val)
                format_cell(cell,color)

            row+=1
        print("="*30)

    wb.save(filename)    


# In[ ]:





# Make some end-game states

# In[74]:


states=[]

state=Board(3,3)
state.pieces=[".","X","O"]  # symbols for empty, player 1, and player 2
state.board=[1,0,2,1,0,2,0,0,0]
print(state)
states.append(state)

state=Board(3,3)
state.pieces=[".","X","O"]  # symbols for empty, player 1, and player 2
state.board=[1,2,2,1,0,2,0,0,1]
print(state)
states.append(state)

state=Board(3,3)
state.pieces=[".","X","O"]  # symbols for empty, player 1, and player 2
state.board=[1,2,2,1,2,2,0,1,1]
print(state)
states.append(state)


# make the table

# In[75]:


T=Table()
player=1

for state in states:
    if state not in T:
        actions=valid_moves(state,player)
        T[state]=Table()
        for action in actions:
            T[state][action]=2  # initial number of skittles


# In[76]:


T


# Write the excel file

# In[77]:


save_table(T,'State Action Value Table.xlsx')


# or load from a file

# In[78]:


T=LoadTable("skittles TTT endgame32.json")


# In[79]:


T


# In[80]:


save_table(T,'State Action Value Table from skittles TTT endgame32.xlsx')


# In[ ]:




